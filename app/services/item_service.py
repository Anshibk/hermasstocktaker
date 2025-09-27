from __future__ import annotations

from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.category import CategoryGroup, SubCategory
from app.models.item import Item
from app.core.constants import CORE_INVENTORY_GROUPS_SET


EXPECTED_HEADERS = {
    "item name": "item_name",
    "group": "group_name",
    "sub category": "sub_category",
    "unit": "unit",
    "price": "price",
}


def _normalise_string(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _coerce_price(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:  # noqa: BLE001
        raise ValueError("Price must be a valid number") from exc


def _resolve_group(db: Session, name: str) -> CategoryGroup:
    existing = (
        db.query(CategoryGroup)
        .filter(func.lower(CategoryGroup.name) == name.lower())
        .one_or_none()
    )
    if existing:
        return existing
    raise ValueError(
        "Inventory group is not recognised. Please use one of the official group names."
    )


def _resolve_subcategory(db: Session, group: CategoryGroup, name: str | None) -> SubCategory | None:
    if not name:
        return None
    existing = (
        db.query(SubCategory)
        .filter(func.lower(SubCategory.name) == name.lower(), SubCategory.group_id == group.id)
        .one_or_none()
    )
    if existing:
        return existing
    sub = SubCategory(name=name, group_id=group.id)
    db.add(sub)
    db.flush()
    return sub


def import_items(db: Session, payload: bytes, *, original_filename: str = "import.xlsx") -> dict[str, int]:
    try:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True)
    except (InvalidFileException, OSError) as exc:  # noqa: BLE001
        raise ValueError("Upload a valid .xlsx workbook") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The workbook is empty")

    header_row = rows[0]
    column_map: dict[int, str] = {}
    seen_headers: set[str] = set()
    for idx, raw_header in enumerate(header_row):
        header_name = _normalise_string(raw_header).lower()
        if not header_name:
            continue
        if header_name in EXPECTED_HEADERS and header_name not in seen_headers:
            column_map[idx] = EXPECTED_HEADERS[header_name]
            seen_headers.add(header_name)

    if set(EXPECTED_HEADERS.values()) - set(column_map.values()):
        raise ValueError(
            "The first row must contain headers for Item Name, Group, Sub Category, Unit, and Price"
        )

    deduped: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
    skipped = 0

    for row in rows[1:]:
        if row is None:
            continue
        record: dict[str, Any] = {"source_row": row}
        for idx, key in column_map.items():
            record[key] = row[idx] if idx < len(row) else None

        item_name = _normalise_string(record.get("item_name"))
        group_name = _normalise_string(record.get("group_name"))
        unit = _normalise_string(record.get("unit"))
        sub_category = _normalise_string(record.get("sub_category"))

        if not item_name or not group_name or not unit:
            skipped += 1
            continue

        if group_name not in CORE_INVENTORY_GROUPS_SET:
            skipped += 1
            continue

        try:
            price = _coerce_price(record.get("price"))
        except ValueError:
            skipped += 1
            continue

        deduped[item_name.lower()] = {
            "item_name": item_name,
            "group_name": group_name,
            "sub_category": sub_category or None,
            "unit": unit,
            "price": price,
        }

    if not deduped:
        raise ValueError("No valid rows found in the spreadsheet")

    created = 0
    updated = 0

    for payload_row in deduped.values():
        try:
            group = _resolve_group(db, payload_row["group_name"])
        except ValueError as exc:  # noqa: PERF203
            raise ValueError(str(exc)) from exc
        sub = _resolve_subcategory(db, group, payload_row.get("sub_category"))

        existing_item = (
            db.query(Item)
            .filter(func.lower(Item.name) == payload_row["item_name"].lower())
            .one_or_none()
        )

        if existing_item:
            existing_item.unit = payload_row["unit"]
            existing_item.price = payload_row["price"]
            existing_item.category_id = sub.id if sub else None
            updated += 1
            continue

        item = Item(
            name=payload_row["item_name"],
            unit=payload_row["unit"],
            price=payload_row["price"],
            category_id=sub.id if sub else None,
        )
        db.add(item)
        created += 1

    db.commit()

    return {
        "filename": original_filename,
        "created": created,
        "updated": updated,
        "skipped": skipped,
    }
