from __future__ import annotations

import uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterator

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import case, func
from sqlalchemy.orm import Session, aliased

from app.models.category import CategoryGroup, SubCategory
from app.models.entry import Entry
from app.models.item import Item
from app.models.session_inv import InventorySession
from app.models.user import User
from app.models.warehouse import Warehouse


CURRENCY_FORMAT = "₹#,##0.00"

HeaderFill = PatternFill("solid", fgColor="0F172A")
HeaderFont = Font(color="FFFFFF", bold=True)
TitleFont = Font(size=14, bold=True)
SubtitleFont = Font(size=11, color="606C80")
ThinBorder = Border(
    left=Side(style="thin", color="CBD5F5"),
    right=Side(style="thin", color="CBD5F5"),
    top=Side(style="thin", color="CBD5F5"),
    bottom=Side(style="thin", color="CBD5F5"),
)


def _decimal_to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):  # noqa: PERF203 - defensive
        return 0.0


def _slugify(value: str) -> str:
    cleaned = [c.lower() if c.isalnum() else "-" for c in value.strip()]
    slug = "".join(cleaned)
    slug = "-".join(filter(None, slug.split("-")))
    return slug or "dashboard-export"


def _safe_sheet_title(value: str) -> str:
    cleaned = "".join(" " if ch in "[]:*?/\\" else ch for ch in value.strip())
    cleaned = cleaned or "Sheet"
    return cleaned[:31]


def _format_decimal_label(value: float | None, decimals: int = 3) -> str:
    if value is None:
        return "0"
    formatted = f"{value:,.{decimals}f}".rstrip("0").rstrip(".")
    return formatted or "0"


def _format_qty_text(value: float | None, unit: str | None) -> str:
    base = _format_decimal_label(_decimal_to_float(value))
    unit_text = (unit or "").strip()
    return f"{base} {unit_text}".strip()


def _format_date_label(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value.date()
    elif isinstance(value, date):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return ""
        try:
            cleaned = text.replace("Z", "+00:00")
            dt = datetime.fromisoformat(cleaned).date()
        except ValueError:
            parts = text.split("-")
            if len(parts) == 3:
                try:
                    dt = date(int(parts[0]), int(parts[1]), int(parts[2][:2]))
                except ValueError:
                    return text
            else:
                return text
    return dt.strftime("%d/%m/%Y")


def _summarise_qty_strings(pairs: list[tuple[float, str | None]]) -> str:
    totals: dict[str, float] = {}
    order: list[str] = []
    for value, unit in pairs:
        unit_key = (unit or "").strip()
        if unit_key not in totals:
            totals[unit_key] = 0.0
            order.append(unit_key)
        totals[unit_key] += _decimal_to_float(value)
    parts = [
        _format_qty_text(totals[key], key if key else None)
        for key in order
        if totals[key]
    ]
    return " + ".join(parts) if parts else "0"


def _stream_entry_rows(query, chunk_size: int = 500) -> Iterator[dict[str, Any]]:
    stream = (
        query.execution_options(stream_results=True)
        .yield_per(chunk_size)
    )
    for row in stream:
        yield {
            "entry_id": row.entry_id,
            "item_id": row.item_id,
            "username": row.username,
            "item_name": row.item_name,
            "category_name": row.category_name,
            "batch": row.batch,
            "mfg": row.mfg,
            "exp": row.exp,
            "qty": _decimal_to_float(row.qty),
            "unit": row.unit,
            "location": row.location,
            "price": None if row.price is None else _decimal_to_float(row.price),
            "line_value": _decimal_to_float(row.line_value),
            "created_at": row.created_at,
        }


def cards(db: Session, user_ids: list[uuid.UUID] | None) -> list[dict[str, Any]]:
    category_rows = (
        db.query(
            CategoryGroup.id.label("group_id"),
            CategoryGroup.name.label("group_name"),
            func.count(func.distinct(SubCategory.id)).label("categories"),
            func.count(func.distinct(Item.id)).label("items"),
        )
        .outerjoin(SubCategory, SubCategory.group_id == CategoryGroup.id)
        .outerjoin(Item, Item.category_id == SubCategory.id)
        .group_by(CategoryGroup.id, CategoryGroup.name)
        .order_by(CategoryGroup.name)
        .all()
    )

    totals_query = (
        db.query(
            CategoryGroup.id.label("group_id"),
            func.count(func.distinct(Entry.item_id)).label("counted"),
            func.sum(
                Entry.qty * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("total_value"),
        )
        .select_from(CategoryGroup)
        .outerjoin(SubCategory, SubCategory.group_id == CategoryGroup.id)
        .outerjoin(Item, Item.category_id == SubCategory.id)
        .outerjoin(Entry, Entry.item_id == Item.id)
    )
    if user_ids:
        totals_query = totals_query.filter(Entry.user_id.in_(user_ids))
    totals = {
        row.group_id: row
        for row in totals_query.group_by(CategoryGroup.id).all()
    }

    results: list[dict[str, Any]] = []
    for row in category_rows:
        metrics = totals.get(row.group_id)
        counted = 0
        total_value = 0.0
        if metrics:
            counted = int(metrics.counted or 0)
            total_value = _decimal_to_float(metrics.total_value)
        results.append(
            {
                "group_name": row.group_name,
                "categories": int(row.categories or 0),
                "items": int(row.items or 0),
                "counted": counted,
                "total_value": total_value,
            }
        )
    return results


def table(db: Session, user_ids: list[uuid.UUID] | None) -> list[dict[str, Any]]:
    query = (
        db.query(
            Entry.item_id.label("item_id"),
            Item.name.label("item_name"),
            Item.unit.label("unit"),
            SubCategory.name.label("category_name"),
            func.count(func.distinct(Entry.batch)).label("batches"),
            func.count(Entry.id).label("entries_logged"),
            func.sum(Entry.qty).label("total_qty"),
            func.sum(
                Entry.qty * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("total_value"),
        )
        .join(Item, Item.id == Entry.item_id)
        .outerjoin(SubCategory, SubCategory.id == Entry.category_id)
    )
    if user_ids:
        query = query.filter(Entry.user_id.in_(user_ids))
    rows = (
        query.group_by(Entry.item_id, Item.name, Item.unit, SubCategory.name)
        .order_by(Item.name)
        .all()
    )

    data: list[dict[str, Any]] = []
    for row in rows:
        data.append(
            {
                "item_id": row.item_id,
                "item_name": row.item_name,
                "unit": row.unit,
                "category_name": row.category_name,
                "batches": int(row.batches or 0),
                "entries_logged": int(row.entries_logged or 0),
                "total_qty": _decimal_to_float(row.total_qty),
                "total_value": _decimal_to_float(row.total_value),
            }
        )
    return data


def detail(
    db: Session,
    user_ids: list[uuid.UUID] | None,
    item_id: uuid.UUID,
    *,
    limit: int | None = 50,
    offset: int = 0,
) -> dict[str, Any]:
    item_row = (
        db.query(
            Item.id.label("item_id"),
            Item.name.label("item_name"),
            SubCategory.name.label("category_name"),
        )
        .outerjoin(SubCategory, Item.category_id == SubCategory.id)
        .filter(Item.id == item_id)
        .one_or_none()
    )
    if not item_row:
        return {"item": None, "entries": []}

    query = (
        db.query(
            Entry.id.label("entry_id"),
            User.username.label("username"),
            Item.name.label("item_name"),
            SubCategory.name.label("category_name"),
            Item.unit.label("unit"),
            Entry.batch,
            Entry.mfg,
            Entry.exp,
            Entry.qty,
            Warehouse.name.label("location"),
            func.coalesce(Entry.price_at_entry, Item.price, 0).label("price"),
            (
                Entry.qty
                * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("line_value"),
            Entry.created_at,
        )
        .join(User, User.id == Entry.user_id)
        .join(Item, Item.id == Entry.item_id)
        .outerjoin(SubCategory, SubCategory.id == Entry.category_id)
        .join(Warehouse, Warehouse.id == Entry.warehouse_id)
        .filter(Entry.item_id == item_id)
    )
    if user_ids:
        query = query.filter(Entry.user_id.in_(user_ids))
    total = query.count()

    ordered_query = query.order_by(Entry.created_at.desc())
    if limit is not None:
        ordered_query = ordered_query.offset(offset).limit(limit)
    rows = ordered_query.all()

    entries: list[dict[str, Any]] = []
    for row in rows:
        qty = _decimal_to_float(row.qty)
        price = _decimal_to_float(row.price)
        line_value = _decimal_to_float(row.line_value)
        entries.append(
            {
                "entry_id": row.entry_id,
                "username": row.username,
                "item_name": row.item_name,
                "category_name": row.category_name,
                "unit": row.unit,
                "batch": row.batch,
                "qty": qty,
                "mfg": row.mfg,
                "exp": row.exp,
                "location": row.location,
                "price": price,
                "line_value": line_value,
                "created_at": row.created_at,
            }
        )

    effective_limit = limit if limit is not None else total
    effective_offset = offset if limit is not None else 0
    has_next = False if limit is None else (effective_offset + len(entries) < total)

    return {
        "item": {
            "item_id": item_row.item_id,
            "item_name": item_row.item_name,
            "category_name": item_row.category_name,
        },
        "entries": entries,
        "total": total,
        "limit": effective_limit,
        "offset": effective_offset,
        "has_next": has_next,
    }


def _make_cell(
    sheet,
    value: Any,
    *,
    alignment: Alignment | None = None,
    font: Font | None = None,
    fill: PatternFill | None = None,
    number_format: str | None = None,
    border: Border | None = ThinBorder,
):
    cell = WriteOnlyCell(sheet, value=value)
    if alignment:
        cell.alignment = alignment
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    return cell


def _append_row(sheet, cells: list[WriteOnlyCell | Any]) -> None:
    """Append a row to a write-only worksheet without type assumptions.

    The openpyxl ``WriteOnlyCell`` helper is implemented as a factory function in
    recent releases, so ``isinstance`` checks against it raise ``TypeError``.
    We only need to ensure we hand write-only sheets a plain sequence of values
    or cells, so pass the list through unchanged.
    """

    sheet.append(list(cells))


def _append_blank_row(sheet, column_count: int) -> None:
    sheet.append([None] * column_count)


def _append_header(sheet, headers: list[str]) -> None:
    header_cells = [
        _make_cell(
            sheet,
            header,
            alignment=Alignment(horizontal="center", vertical="center"),
            font=HeaderFont,
            fill=HeaderFill,
        )
        for header in headers
    ]
    _append_row(sheet, header_cells)


def _write_title_block(sheet, title: str, generated_label: str | None = None) -> None:
    title_cell = _make_cell(
        sheet,
        title,
        font=TitleFont,
        alignment=Alignment(horizontal="left"),
    )
    _append_row(sheet, [title_cell])
    if generated_label:
        subtitle_cell = _make_cell(
            sheet,
            f"Generated on {generated_label}",
            font=SubtitleFont,
            alignment=Alignment(horizontal="left"),
        )
        _append_row(sheet, [subtitle_cell])


def _append_key_value_row(
    sheet,
    label: str,
    value: Any,
    *,
    number_format: str | None = None,
) -> None:
    _append_row(
        sheet,
        [
            _make_cell(
                sheet,
                label,
                font=Font(size=11, bold=True, color="334155"),
                fill=PatternFill("solid", fgColor="E2E8F0"),
                alignment=Alignment(horizontal="left"),
            ),
            _make_cell(
                sheet,
                value,
                alignment=Alignment(horizontal="left"),
                number_format=number_format,
            ),
        ],
    )


def export_detail(
    db: Session,
    user_ids: list[uuid.UUID] | None,
    item_id: uuid.UUID,
) -> tuple[BytesIO, str]:
    item_row = (
        db.query(
            Item.id.label("item_id"),
            Item.name.label("item_name"),
            SubCategory.name.label("category_name"),
        )
        .outerjoin(SubCategory, Item.category_id == SubCategory.id)
        .filter(Item.id == item_id)
        .one_or_none()
    )
    if not item_row:
        raise ValueError("Item not found")

    totals_query = (
        db.query(
            func.count(func.distinct(Entry.batch)).label("batches"),
            func.count(Entry.id).label("entries_logged"),
            func.sum(Entry.qty).label("total_qty"),
            func.sum(
                Entry.qty * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("total_value"),
            func.max(Item.unit).label("unit"),
        )
        .join(Item, Item.id == Entry.item_id)
        .filter(Entry.item_id == item_id)
    )
    if user_ids:
        totals_query = totals_query.filter(Entry.user_id.in_(user_ids))
    totals = totals_query.one()

    item_name = item_row.item_name or "Inventory Item"
    category_name = item_row.category_name or "Uncategorised"

    total_qty = _decimal_to_float(totals.total_qty)
    total_value = _decimal_to_float(totals.total_value)
    total_batches = int(totals.batches or 0)
    entries_logged = int(totals.entries_logged or 0)
    unit = totals.unit

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=_safe_sheet_title(f"{item_name}_Summary"))

    timestamp_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")

    title_row = [
        _make_cell(
            sheet,
            item_name,
            font=Font(size=16, bold=True, color="1F2937"),
            alignment=Alignment(horizontal="left"),
        )
    ]
    _append_row(sheet, title_row)
    _append_row(sheet, [
        _make_cell(
            sheet,
            f"Category: {category_name or '—'}",
            font=SubtitleFont,
            alignment=Alignment(horizontal="left"),
        )
    ])
    _append_row(sheet, [
        _make_cell(
            sheet,
            f"Exported On: {timestamp_label}",
            font=Font(size=10, color="64748B"),
            alignment=Alignment(horizontal="left"),
        )
    ])

    qty_text = _format_qty_text(total_qty, unit)
    summary_pairs = [
        ("Total Qty Counted", qty_text, None),
        ("Total Inventory Value", total_value, CURRENCY_FORMAT),
        ("Batches", total_batches, "0"),
        ("Entries logged", entries_logged, "0"),
    ]

    for label, value, number_format in summary_pairs:
        _append_key_value_row(
            sheet,
            label,
            value,
            number_format=number_format,
        )

    headers = [
        "#",
        "Entry Date",
        "User",
        "Batch",
        "Mfg",
        "Exp",
        "Quantity",
        "Location",
        "Price",
        "Line Value",
    ]
    _append_blank_row(sheet, len(headers))
    _append_header(sheet, headers)

    entry_category = aliased(SubCategory)
    item_category_detail = aliased(SubCategory)
    detail_query = (
        db.query(
            Entry.id.label("entry_id"),
            Entry.item_id.label("item_id"),
            User.username.label("username"),
            Item.name.label("item_name"),
            func.coalesce(entry_category.name, item_category_detail.name).label(
                "category_name"
            ),
            Entry.batch,
            Entry.mfg,
            Entry.exp,
            Entry.qty,
            Warehouse.name.label("location"),
            func.coalesce(Entry.price_at_entry, Item.price).label("price"),
            (
                Entry.qty
                * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("line_value"),
            Item.unit.label("unit"),
            Entry.created_at.label("created_at"),
        )
        .join(User, User.id == Entry.user_id)
        .join(Item, Item.id == Entry.item_id)
        .outerjoin(entry_category, entry_category.id == Entry.category_id)
        .outerjoin(item_category_detail, item_category_detail.id == Item.category_id)
        .join(Warehouse, Warehouse.id == Entry.warehouse_id)
        .filter(Entry.item_id == item_id)
        .order_by(Entry.created_at)
    )
    if user_ids:
        detail_query = detail_query.filter(Entry.user_id.in_(user_ids))

    for index, entry in enumerate(_stream_entry_rows(detail_query), start=1):
        price = entry.get("price")
        price_format = CURRENCY_FORMAT if price is not None else None
        _append_row(
            sheet,
            [
                _make_cell(sheet, index, alignment=Alignment(horizontal="center")),
                _make_cell(
                    sheet,
                    _format_date_label(entry.get("created_at")),
                    alignment=Alignment(horizontal="center"),
                ),
                _make_cell(sheet, entry.get("username"), alignment=Alignment(horizontal="left")),
                _make_cell(sheet, entry.get("batch"), alignment=Alignment(horizontal="left")),
                _make_cell(sheet, entry.get("mfg"), alignment=Alignment(horizontal="center")),
                _make_cell(sheet, entry.get("exp"), alignment=Alignment(horizontal="center")),
                _make_cell(
                    sheet,
                    _format_qty_text(entry.get("qty"), entry.get("unit")),
                    alignment=Alignment(horizontal="right"),
                ),
                _make_cell(sheet, entry.get("location"), alignment=Alignment(horizontal="left")),
                _make_cell(
                    sheet,
                    price,
                    alignment=Alignment(horizontal="right"),
                    number_format=price_format,
                ),
                _make_cell(
                    sheet,
                    entry.get("line_value"),
                    alignment=Alignment(horizontal="right"),
                    number_format=CURRENCY_FORMAT,
                ),
            ],
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"{_slugify(str(item_name))}_summary.xlsx"
    return stream, filename

def export_dashboard(
    db: Session,
    user_ids: list[uuid.UUID] | None,
    *,
    include_master_items: bool,
) -> tuple[BytesIO, str]:
    session_rows = (
        db.query(InventorySession.id)
        .filter(InventorySession.status == "active")
        .all()
    )
    session_ids = [row.id for row in session_rows]

    aggregated_base = (
        db.query(
            Entry.item_id.label("item_id"),
            func.count(func.distinct(Entry.batch)).label("batches"),
            func.count(Entry.id).label("entries_logged"),
            func.sum(Entry.qty).label("total_qty"),
            func.sum(
                Entry.qty * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("total_value"),
        )
        .join(Item, Item.id == Entry.item_id)
    )
    if user_ids:
        aggregated_base = aggregated_base.filter(Entry.user_id.in_(user_ids))
    if session_ids:
        aggregated_base = aggregated_base.filter(Entry.session_id.in_(session_ids))
    aggregated_subq = aggregated_base.group_by(Entry.item_id).subquery()

    item_category = aliased(SubCategory)
    item_group = aliased(CategoryGroup)

    items_query = (
        db.query(
            Item.id.label("item_id"),
            Item.name.label("item_name"),
            Item.unit.label("unit"),
            item_category.name.label("category_name"),
            item_group.name.label("group_name"),
            aggregated_subq.c.batches,
            aggregated_subq.c.entries_logged,
            aggregated_subq.c.total_qty,
            aggregated_subq.c.total_value,
        )
        .outerjoin(item_category, item_category.id == Item.category_id)
        .outerjoin(item_group, item_group.id == item_category.group_id)
        .outerjoin(aggregated_subq, aggregated_subq.c.item_id == Item.id)
        .order_by(
            case((aggregated_subq.c.entries_logged.is_(None), 1), else_=0),
            func.lower(Item.name),
        )
    )

    generated_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
    download_date = datetime.utcnow().strftime("%d-%b-%Y")
    workbook = Workbook(write_only=True)

    items_headers = [
        "#",
        "Item Name",
        "Category",
        "Group",
        "Batches",
        "Entries Logged",
        "Total Quantity",
        "Total Value",
        "Unit",
    ]

    if include_master_items:
        items_sheet = workbook.create_sheet(title="Master Items")
        _write_title_block(items_sheet, "Master Item Inventory Report", generated_label)
        _append_header(items_sheet, items_headers)

        item_rows = items_query.execution_options(stream_results=True).yield_per(500)
        for index, row in enumerate(item_rows, start=1):
            batches = int(row.batches or 0)
            entries_logged = int(row.entries_logged or 0)
            total_qty = _format_qty_text(row.total_qty, row.unit)
            _append_row(
                items_sheet,
                [
                    _make_cell(items_sheet, index, alignment=Alignment(horizontal="center")),
                    _make_cell(items_sheet, row.item_name, alignment=Alignment(horizontal="left")),
                    _make_cell(items_sheet, row.category_name or "—", alignment=Alignment(horizontal="left")),
                    _make_cell(items_sheet, row.group_name or "—", alignment=Alignment(horizontal="left")),
                    _make_cell(items_sheet, batches, alignment=Alignment(horizontal="center"), number_format="0"),
                    _make_cell(items_sheet, entries_logged, alignment=Alignment(horizontal="center"), number_format="0"),
                    _make_cell(items_sheet, total_qty, alignment=Alignment(horizontal="right")),
                    _make_cell(
                        items_sheet,
                        _decimal_to_float(row.total_value),
                        alignment=Alignment(horizontal="right"),
                        number_format=CURRENCY_FORMAT,
                    ),
                    _make_cell(items_sheet, row.unit, alignment=Alignment(horizontal="left")),
                ],
            )

        filename = f"export_with_master_items_{download_date}.xlsx"
    else:
        entry_category = aliased(SubCategory)
        item_category_detail = aliased(SubCategory)
        detail_query = (
            db.query(
                Entry.id.label("entry_id"),
                Entry.item_id.label("item_id"),
                User.username.label("username"),
                Item.name.label("item_name"),
                func.coalesce(entry_category.name, item_category_detail.name).label(
                    "category_name"
                ),
                Entry.batch,
                Entry.mfg,
                Entry.exp,
                Entry.qty,
                Warehouse.name.label("location"),
                func.coalesce(Entry.price_at_entry, Item.price).label("price"),
                (
                    Entry.qty
                    * func.coalesce(Entry.price_at_entry, Item.price, 0)
                ).label("line_value"),
                Item.unit.label("unit"),
                Entry.created_at.label("created_at"),
            )
            .join(User, User.id == Entry.user_id)
            .join(Item, Item.id == Entry.item_id)
            .outerjoin(entry_category, entry_category.id == Entry.category_id)
            .outerjoin(item_category_detail, item_category_detail.id == Item.category_id)
            .join(Warehouse, Warehouse.id == Entry.warehouse_id)
        )
        if user_ids:
            detail_query = detail_query.filter(Entry.user_id.in_(user_ids))
        if session_ids:
            detail_query = detail_query.filter(Entry.session_id.in_(session_ids))
        detail_query = detail_query.order_by(Item.name, Entry.created_at)

        valued_sheet = workbook.create_sheet(title="Valued Items")
        _write_title_block(valued_sheet, "Valued Item Report", generated_label)
        _append_header(valued_sheet, items_headers)

        valued_query = items_query.filter(aggregated_subq.c.entries_logged.isnot(None))
        valued_rows = valued_query.execution_options(stream_results=True).yield_per(500)

        for index, row in enumerate(valued_rows, start=1):
            batches = int(row.batches or 0)
            entries_logged = int(row.entries_logged or 0)
            total_qty = _format_qty_text(row.total_qty, row.unit)
            _append_row(
                valued_sheet,
                [
                    _make_cell(valued_sheet, index, alignment=Alignment(horizontal="center")),
                    _make_cell(valued_sheet, row.item_name, alignment=Alignment(horizontal="left")),
                    _make_cell(valued_sheet, row.category_name or "—", alignment=Alignment(horizontal="left")),
                    _make_cell(valued_sheet, row.group_name or "—", alignment=Alignment(horizontal="left")),
                    _make_cell(valued_sheet, batches, alignment=Alignment(horizontal="center"), number_format="0"),
                    _make_cell(valued_sheet, entries_logged, alignment=Alignment(horizontal="center"), number_format="0"),
                    _make_cell(valued_sheet, total_qty, alignment=Alignment(horizontal="right")),
                    _make_cell(
                        valued_sheet,
                        _decimal_to_float(row.total_value),
                        alignment=Alignment(horizontal="right"),
                        number_format=CURRENCY_FORMAT,
                    ),
                    _make_cell(valued_sheet, row.unit, alignment=Alignment(horizontal="left")),
                ],
            )

        entries_sheet = workbook.create_sheet(title="Valued Entries")
        _write_title_block(entries_sheet, "Valued Logged Entries", generated_label)
        entry_headers = [
            "#",
            "Entry Date",
            "User",
            "Item Name",
            "Category",
            "Batch",
            "Mfg",
            "Exp",
            "Quantity",
            "Location",
            "Price",
            "Line Value",
        ]
        _append_header(entries_sheet, entry_headers)

        filtered_detail = detail_query.filter(Entry.id.isnot(None))
        for index, entry in enumerate(_stream_entry_rows(filtered_detail), start=1):
            price = entry.get("price")
            price_format = CURRENCY_FORMAT if price is not None else None
            _append_row(
                entries_sheet,
                [
                    _make_cell(entries_sheet, index, alignment=Alignment(horizontal="center")),
                    _make_cell(
                        entries_sheet,
                        _format_date_label(entry.get("created_at")),
                        alignment=Alignment(horizontal="center"),
                    ),
                    _make_cell(entries_sheet, entry.get("username"), alignment=Alignment(horizontal="left")),
                    _make_cell(entries_sheet, entry.get("item_name"), alignment=Alignment(horizontal="left")),
                    _make_cell(entries_sheet, entry.get("category_name") or "—", alignment=Alignment(horizontal="left")),
                    _make_cell(entries_sheet, entry.get("batch") or "—", alignment=Alignment(horizontal="left")),
                    _make_cell(entries_sheet, entry.get("mfg"), alignment=Alignment(horizontal="center")),
                    _make_cell(entries_sheet, entry.get("exp"), alignment=Alignment(horizontal="center")),
                    _make_cell(
                        entries_sheet,
                        _format_qty_text(entry.get("qty"), entry.get("unit")),
                        alignment=Alignment(horizontal="right"),
                    ),
                    _make_cell(entries_sheet, entry.get("location") or "—", alignment=Alignment(horizontal="left")),
                    _make_cell(
                        entries_sheet,
                        price,
                        alignment=Alignment(horizontal="right"),
                        number_format=price_format,
                    ),
                    _make_cell(
                        entries_sheet,
                        entry.get("line_value"),
                        alignment=Alignment(horizontal="right"),
                        number_format=CURRENCY_FORMAT,
                    ),
                ],
            )

        filename = f"export_valued_items_{download_date}.xlsx"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream, filename
